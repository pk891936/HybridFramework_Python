import openpyxl

def getRowCount(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return (sheet.max_row)
def getColumnCount(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return (sheet.max_col)
def readData(file,sheetName,rownum,colnum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.cell(row=rownum, column=colnum).value
def writeData(file,sheetName,rownum,colnum,data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    data = sheet.cell(row=rownum, column=colnum).value
    workbook.save(file)

