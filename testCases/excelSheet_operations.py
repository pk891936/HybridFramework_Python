import openpyxl

workbook = openpyxl.load_workbook("C://Users//Praveen//PycharmProjects//Hybrid_Framework//testData//testdata.xlsx")
sheet = workbook["Sheet2"]
row  = sheet.max_row
col = sheet.max_column
print("col =",col)

count = 0
for i in range(2,row):
    if int(sheet.cell(row=i, column=2).value )>45:
        count = count +1
sheet = workbook["Sheet2"]
sheet.cell(row=row+1, column=1).value = "total>45"
sheet.cell(row=row+1, column=2).value = count
workbook.save("C://Users//Praveen//PycharmProjects//Hybrid_Framework//testData//testdata.xlsx")
print(count)